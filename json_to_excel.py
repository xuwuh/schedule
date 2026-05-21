"""Экспорт готового расписания JSON в Excel.

Файл вынесен из функций app.py исходной сборки:
- lessons_by_group
- sorted_lessons
- autosize
- write_schedule_sheet
- export_schedule_excel

Ожидаемый формат входного JSON:
{
    "lessons": [
        {
            "день_недели": "Понедельник",
            "время": "09:00 – 10:30",
            "группа": "11",
            "предмет": "математика",
            "тип": "Лекция",
            "преподаватель": "Иванов Иван Иванович",
            "аудитория": "101"
        }
    ]
}
"""

import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_FILE = Path("schedule.json")
DEFAULT_OUTPUT_FILE = Path("schedule_export.xlsx")

DAYS = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]


def load_schedule(input_path):
    with open(input_path, "r", encoding="utf-8") as file:
        return json.load(file)


def lessons_by_group(schedule_data):
    lessons = schedule_data.get("lessons", []) if isinstance(schedule_data, dict) else []
    groups = defaultdict(list)

    for lesson in lessons:
        raw_groups = str(lesson.get("группа", "")).split(",")
        for group in [group.strip() for group in raw_groups if group.strip()]:
            groups[group].append(lesson)

    return dict(sorted(groups.items(), key=lambda item: item[0]))


def sorted_lessons(lessons):
    day_index = {day: index for index, day in enumerate(DAYS)}
    return sorted(
        lessons,
        key=lambda item: (
            day_index.get(item.get("день_недели", ""), 99),
            str(item.get("время", "")),
        ),
    )


def autosize(sheet):
    for column_cells in sheet.columns:
        max_len = 10
        letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))

        sheet.column_dimensions[letter].width = min(max_len + 3, 42)


def write_schedule_sheet(workbook, title, lessons):
    sheet = workbook.create_sheet(title[:31])

    sheet.append([
        "День недели",
        "Время",
        "Группа",
        "Предмет",
        "Тип",
        "Преподаватель",
        "Аудитория",
    ])

    fill = PatternFill("solid", fgColor="3657FF")
    font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for lesson in sorted_lessons(lessons):
        sheet.append([
            lesson.get("день_недели", ""),
            lesson.get("время", ""),
            lesson.get("группа", ""),
            lesson.get("предмет", ""),
            lesson.get("тип", ""),
            lesson.get("преподаватель", ""),
            lesson.get("аудитория", ""),
        ])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    autosize(sheet)


def export_schedule_excel(schedule_data, output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    lessons = schedule_data.get("lessons", [])
    write_schedule_sheet(workbook, "Общее расписание", lessons)

    for group, group_lessons in lessons_by_group(schedule_data).items():
        write_schedule_sheet(workbook, f"Группа {group}", group_lessons)

    workbook.save(output_path)


def convert_json_to_excel(input_file=DEFAULT_INPUT_FILE, output_file=DEFAULT_OUTPUT_FILE):
    input_file = Path(input_file)
    output_file = Path(output_file)

    schedule_data = load_schedule(input_file)
    export_schedule_excel(schedule_data, output_file)

    return output_file, len(schedule_data.get("lessons", []))


def parse_args():
    parser = argparse.ArgumentParser(
        description="Конвертирует готовое расписание schedule.json в Excel .xlsx."
    )
    parser.add_argument("input", nargs="?", default=str(DEFAULT_INPUT_FILE))
    parser.add_argument("output", nargs="?", default=str(DEFAULT_OUTPUT_FILE))
    return parser.parse_args()


def main():
    args = parse_args()
    output_file, lesson_count = convert_json_to_excel(args.input, args.output)
    print(f"Excel создан: {output_file}")
    print(f"Занятий экспортировано: {lesson_count}")


if __name__ == "__main__":
    main()
