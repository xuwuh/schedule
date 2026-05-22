import json
import tempfile
import importlib.util
from pathlib import Path

from flask import Flask, jsonify, request, send_file, send_from_directory

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


def load_module(alias, filename):
    path = BASE_DIR / filename
    spec = importlib.util.spec_from_file_location(alias, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

checker = load_module("checker_original", "checker.py")
dynamic_scheduler = load_module("dynamic_scheduler_original", "dynamic_scheduler.py")
excel_to_json = load_module("excel_to_json_original", "excel_to_json.py")
json_to_calendar = load_module("json_to_calendar_original", "json_to_calendar.py")
json_to_excel = load_module("json_to_excel_original", "json_to_excel.py")
json_to_pdf = load_module("json_to_pdf_original", "json_to_pdf.py")
shebule = load_module("shebule_original", "shebule.py")

app = Flask(__name__)


def error_response(error, code=400):
    return jsonify({"error": str(error)}), code


def save_upload_to_temp(field="file", suffix=".xlsx"):
    file = request.files.get(field)
    if not file:
        raise ValueError("Файл не выбран.")
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix, dir=UPLOAD_DIR)
    temp.close()
    file.save(temp.name)
    return Path(temp.name)


def make_schedule_template(path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Все занятия"
    headers = ["День недели", "Время", "Группа", "Предмет", "Тип", "Преподаватель", "Аудитория"]
    sheet.append(headers)
    sheet.append(["Понедельник", "09:00 – 10:30", "1011", "Математика", "Лекция", "Иванов Иван Иванович", "101"])

    fill = PatternFill("solid", fgColor="4A509A")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(len(header) + 8, 18)
    sheet.freeze_panes = "A2"
    workbook.save(path)


def normalize_header(value):
    """Нормализует заголовки Excel, чтобы принимать разные варианты названий."""
    return str(value or "").strip().lower().replace("ё", "е").replace("_", " ")


def get_row_value(row, aliases, default=""):
    """Берёт значение из строки Excel по одному из возможных названий колонки."""
    for alias in aliases:
        key = normalize_header(alias)
        if key in row and row[key] not in (None, ""):
            return row[key]
    return default


def find_schedule_sheet(workbook):
    """Ищет лист именно с готовым расписанием, а не с исходными данными."""
    schedule_sheet_names = {
        "все занятия",
        "расписание",
        "готовое расписание",
        "schedule",
        "lessons",
    }

    for sheet_name in workbook.sheetnames:
        normalized = normalize_header(sheet_name)
        if normalized in schedule_sheet_names:
            return workbook[sheet_name]

    # Если точного названия нет — пробуем найти лист по заголовкам.
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = {normalize_header(value) for value in rows[0] if value}
        needed = {"время", "группа", "предмет", "преподаватель"}

        if needed.issubset(headers):
            return sheet

    return None


def parse_schedule_excel(path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)

    sheet = None

    for sheet_name in workbook.sheetnames:
        name = sheet_name.strip().lower()

        if name in [
            "общее расписание",
            "все занятия",
            "расписание",
            "schedule",
            "lessons"
        ]:
            sheet = workbook[sheet_name]
            break

    if sheet is None:
        raise ValueError(
            "Не найден лист с готовым расписанием. "
            "Нужен лист: 'Общее расписание' или 'Все занятия'."
        )

    headers = [
        str(cell.value or "").strip().lower().replace(" ", "_")
        for cell in sheet[1]
    ]

    lessons = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in row):
            continue

        item = dict(zip(headers, row))

        lesson = {
            "_id": f"lesson-{row_index}",
            "день_недели": str(item.get("день_недели") or item.get("день") or "").strip(),
            "время": str(item.get("время") or "").strip(),
            "группа": str(item.get("группа") or "").strip(),
            "предмет": str(item.get("предмет") or "").strip(),
            "тип": str(item.get("тип") or "").strip(),
            "преподаватель": str(item.get("преподаватель") or "").strip(),
            "аудитория": str(item.get("аудитория") or "").strip()
        }

        if lesson["день_недели"] and lesson["время"]:
            lessons.append(lesson)

    return {
        "lessons": lessons,
        "unplaced_lessons": []
    }


@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/styles.css")
def styles():
    return send_from_directory(BASE_DIR, "styles.css")


@app.route("/script.js")
def script():
    return send_from_directory(BASE_DIR, "script.js")


@app.post("/api/upload-excel")
def upload_excel():
    try:
        path = save_upload_to_temp(suffix=".xlsx")
        output = path.with_suffix(".json")
        data = excel_to_json.convert_excel_to_json(path, output)
        return jsonify({"data": data})
    except Exception as error:
        return error_response(error)


@app.post("/api/upload-schedule")
def upload_schedule():
    try:
        path = save_upload_to_temp(suffix=".xlsx")

        data = parse_schedule_excel(path)
        normalized = dynamic_scheduler.normalize_schedule(data)

        return jsonify({"data": normalized})

    except Exception as error:
        return error_response(error)


@app.post("/api/generate")
def generate():
    try:
        data = request.get_json(force=True)
        result = shebule.schedule(data)
        return jsonify(result)
    except Exception as error:
        return error_response(error)


@app.post("/api/validate")
def validate():
    try:
        schedule_data = request.get_json(force=True)
        errors = checker.validator(schedule_data)
        unplaced = schedule_data.get("unplaced_lessons", []) if isinstance(schedule_data, dict) else []
        return jsonify({"errors": errors, "count": len(errors), "unplaced_lessons": unplaced})
    except Exception as error:
        return error_response(error)


@app.post("/api/dynamic/options")
def dynamic_options():
    try:
        payload = request.get_json(force=True)
        return jsonify(dynamic_scheduler.extract_options(payload.get("source_data", {}), payload.get("schedule", {})))
    except Exception as error:
        return error_response(error)


@app.post("/api/dynamic/apply")
def dynamic_apply():
    try:
        payload = request.get_json(force=True)
        result = dynamic_scheduler.apply_dynamic_change(
            payload.get("schedule", {}),
            payload.get("source_data", {}),
            payload.get("change", {}),
        )
        return jsonify(result)
    except Exception as error:
        return error_response(error)


def write_temp_schedule(schedule_data):
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".json", dir=OUTPUT_DIR, mode="w", encoding="utf-8")
    with temp:
        json.dump(schedule_data, temp, ensure_ascii=False, indent=4)
    return Path(temp.name)


@app.post("/api/excel")
def export_excel():
    try:
        payload = request.get_json(force=True)
        input_path = write_temp_schedule(payload.get("schedule", {}))
        output_path = input_path.with_suffix(".xlsx")
        json_to_excel.convert_json_to_excel(input_path, output_path)
        return send_file(output_path, as_attachment=True, download_name="schedule_export.xlsx")
    except Exception as error:
        return error_response(error)


@app.post("/api/pdf")
def export_pdf():
    try:
        payload = request.get_json(force=True)
        input_path = write_temp_schedule(payload.get("schedule", {}))
        output_path = input_path.with_suffix(".pdf")
        json_to_pdf.convert_json_to_pdf(input_path, output_path)
        return send_file(output_path, as_attachment=True, download_name="schedule_export.pdf")
    except Exception as error:
        return error_response(error)


@app.post("/api/calendar")
def export_calendar():
    try:
        payload = request.get_json(force=True)
        input_path = write_temp_schedule(payload.get("schedule", {}))
        output_path = input_path.with_suffix(".ics")
        repeat_weeks = int(payload.get("repeat_weeks") or 1)
        json_to_calendar.convert_json_to_calendar(input_file=input_path, output_file=output_path, repeat_weeks=repeat_weeks)
        return send_file(output_path, as_attachment=True, download_name="schedule.ics")
    except Exception as error:
        return error_response(error)


@app.get("/download/template")
def download_data_template():
    try:
        path = OUTPUT_DIR / "schedule_data_template.xlsx"
        excel_to_json.create_template(path)
        return send_file(path, as_attachment=True, download_name="schedule_data_template.xlsx")
    except Exception as error:
        return error_response(error)


@app.get("/download/schedule-template")
def download_schedule_template():
    try:
        path = OUTPUT_DIR / "schedule_template.xlsx"
        make_schedule_template(path)
        return send_file(path, as_attachment=True, download_name="schedule_template.xlsx")
    except Exception as error:
        return error_response(error)


if __name__ == "__main__":
    app.run(debug=True)
