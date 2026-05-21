import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_FILE = Path("schedule_data.xlsx")
DEFAULT_OUTPUT_FILE = Path("schedule_data.json")
DEFAULT_TEMPLATE_FILE = Path("schedule_data_template.xlsx")

SOURCE_SHEETS = {
    "аудитории": "аудитории",
    "auditoriums": "аудитории",
    "преподаватели": "преподаватели",
    "teachers": "преподаватели",
    "предметы": "предметы",
    "subjects": "предметы",
    "группы": "группы",
    "groups": "группы",
}

SCHEDULE_SHEETS = {"все занятия", "schedule", "lessons"}


def normalize_text(value):
    return str(value).strip() if value is not None else ""


def normalize_header(value):
    return normalize_text(value).lower().replace("ё", "е")


def parse_int(value, field_name, row_number):
    if value is None or value == "":
        raise ValueError(f"строка {row_number}: поле '{field_name}' не заполнено")
    try:
        return int(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"строка {row_number}: поле '{field_name}' должно быть числом") from error


def parse_optional_int(value, default=0):
    if value is None or value == "":
        return default
    return int(value)


def parse_list(value, *, item_type=str):
    if value is None or value == "":
        return []

    if isinstance(value, (int, float)):
        values = [value]
    else:
        text = str(value).replace(";", ",")
        values = [part.strip() for part in text.split(",")]

    result = []
    for item in values:
        if item == "":
            continue
        if item_type is int:
            result.append(int(float(item)))
        else:
            result.append(str(item).strip())
    return result


def parse_lesson_type(value):
    text = normalize_text(value).lower()
    if text in {"1", "лекция", "lecture"}:
        return 1
    if text in {"0", "семинар", "seminar"}:
        return 0
    raise ValueError(f"неизвестный тип занятия: {value}")

def parse_subject_names(value, subject_map):
    if value is None or value == "":
        return []
    result = []
    parts = str(value).replace(";", ",").split(",")
    for part in parts:
        key = part.strip().lower().replace("ё", "е")
        if key in subject_map:
            result.append(subject_map[key])
        else:
            raise ValueError(f"Предмет не найден в листе 'предметы': {part}")
    return result

def sheet_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(value) for value in rows[0]]
    result = []

    for row_index, row in enumerate(rows[1:], start=2):
        if all(value is None or normalize_text(value) == "" for value in row):
            continue
        item = {}
        for header, value in zip(headers, row):
            if header:
                item[header] = value
        item["_row_number"] = row_index
        result.append(item)

    return result


def get_value(row, aliases, default=None):
    for alias in aliases:
        key = normalize_header(alias)
        if key in row and row[key] is not None:
            return row[key]
    return default


def convert_auditoriums(rows):
    auditoriums = []
    for row in rows:
        row_number = row["_row_number"]
        aud_type = parse_list(get_value(row, ["тип аудитории", "тип", "условия"], "-"))
        auditoriums.append({
            "id": parse_int(get_value(row, ["id", "ид"]), "id", row_number),
            "номер": parse_int(get_value(row, ["номер", "номер аудитории"]), "номер", row_number),
            "вместимость": parse_int(get_value(row, ["вместимость", "мест"]), "вместимость", row_number),
            "тип аудитории": aud_type or ["-"],
        })
    return auditoriums


def convert_teachers(rows, subject_map):
    teachers = []
    for row in rows:
        row_number = row["_row_number"]
        teachers.append({
            "id": parse_int(get_value(row, ["id", "ид"]), "id", row_number),
            "ФИО": normalize_text(get_value(row, ["фио", "преподаватель", "имя"])),
            "занятость в неделю": parse_int(
                get_value(row, ["занятость в неделю", "нагрузка", "часов в неделю"]),
                "занятость в неделю",
                row_number,
            ),
            "преподает предметы (id)": parse_subject_names(
                get_value(row, ["преподает предметы (id)", "предметы id", "предметы"]),
                subject_map
            ),
        })

    return teachers


def convert_subjects(rows):
    subjects = []
    for row in rows:
        row_number = row["_row_number"]
        conditions = parse_list(
            get_value(row, ["дополнительны условия", "дополнительные условия", "условия"], "-")
        )
        subjects.append({
            "id": parse_int(get_value(row, ["id", "ид"]), "id", row_number),
            "название": normalize_text(get_value(row, ["название", "предмет"])),
            "часы в неделю": parse_int(
                get_value(row, ["часы в неделю", "часы", "нагрузка"]),
                "часы в неделю",
                row_number,
            ),
            "тип занятия": parse_lesson_type(get_value(row, ["тип занятия", "тип"])),
            "длительность": parse_optional_int(get_value(row, ["длительность", "минуты"]), 45),
            "дополнительны условия": conditions or ["-"],
        })
    return subjects


def convert_groups(rows, subject_map):
    groups = []
    for row in rows:
        row_number = row["_row_number"]
        groups.append({
            "id": parse_int(get_value(row, ["id", "ид"]), "id", row_number),
            "номер": parse_int(get_value(row, ["номер", "номер группы", "группа"]), "номер", row_number),
            "кол-во человек": parse_int(
                get_value(row, ["кол-во человек", "количество человек", "студентов"]),
                "кол-во человек",
                row_number,
            ),
            "предметы (id)": parse_subject_names(
                get_value(row, ["предметы (id)", "предметы id", "предметы"]),
                subject_map
            ),
        })

    return groups

def build_subject_map(subjects):
    subject_map = {}
    for subject in subjects:
        lesson_type = (
            "лекция"
            if subject["тип занятия"] == 1
            else "семинар"
        )
        key = f'{subject["название"].lower()} ({lesson_type})'.replace("ё", "е")
        subject_map[key] = subject["id"]
    return subject_map


def find_source_sheets(workbook):
    found = {}
    for sheet_name in workbook.sheetnames:
        normalized = normalize_header(sheet_name)
        if normalized in SOURCE_SHEETS:
            found[SOURCE_SHEETS[normalized]] = workbook[sheet_name]
    return found


def convert_source_workbook(workbook):
    sheets = find_source_sheets(workbook)
    required = ["аудитории", "преподаватели", "предметы", "группы"]
    missing = [name for name in required if name not in sheets]
    if missing:
        raise ValueError(f"В Excel-файле не найдены обязательные листы: {', '.join(missing)}")

    subjects = convert_subjects(sheet_rows(sheets["предметы"]))
    subject_map = build_subject_map(subjects)
    teachers = convert_teachers(
        sheet_rows(sheets["преподаватели"]),
        subject_map
    )

    groups = convert_groups(
        sheet_rows(sheets["группы"]),
        subject_map
    )

    return {
        "аудитории": convert_auditoriums(sheet_rows(sheets["аудитории"])),
        "преподаватели": teachers,
        "предметы": subjects,
        "группы": groups,
    }


def convert_schedule_workbook(workbook):
    sheet = None
    for sheet_name in workbook.sheetnames:
        if normalize_header(sheet_name) in SCHEDULE_SHEETS:
            sheet = workbook[sheet_name]
            break

    if sheet is None:
        return None

    lessons = []
    for row in sheet_rows(sheet):
        lessons.append({
            "день_недели": normalize_text(get_value(row, ["день недели", "день"])),
            "время": normalize_text(get_value(row, ["время"])),
            "группа": normalize_text(get_value(row, ["группа"])),
            "предмет": normalize_text(get_value(row, ["предмет"])),
            "тип": normalize_text(get_value(row, ["тип"])),
            "преподаватель": normalize_text(get_value(row, ["преподаватель"])),
            "аудитория": normalize_text(get_value(row, ["аудитория"])),
        })

    return {"lessons": lessons, "unplaced_lessons": []}


def convert_excel_to_json(input_file, output_file):
    workbook = load_workbook(input_file, data_only=True)

    source_sheets = find_source_sheets(workbook)
    if source_sheets:
        data = convert_source_workbook(workbook)
    else:
        data = convert_schedule_workbook(workbook)
        if data is None:
            raise ValueError(
                "Не удалось определить формат Excel-файла. "
                "Используйте листы 'аудитории', 'преподаватели', 'предметы', 'группы' "
                "или лист 'Все занятия'."
            )

    with open(output_file, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

    return data


def write_headers(sheet, headers):
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="4A509A")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, header in enumerate(headers, start=1):
        width = max(len(header) + 4, 14)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.freeze_panes = "A2"


def append_example_rows(workbook):
    sheet = workbook["аудитории"]
    sheet.append([1, 101, 60, "экран"])
    sheet.append([2, 102, 30, "тех, экран"])

    sheet = workbook["преподаватели"]
    sheet.append([1, "Иванов Иван Иванович", 18, "1, 2"])
    sheet.append([2, "Петрова Анна Сергеевна", 20, "3"])

    sheet = workbook["предметы"]
    sheet.append([1, "Математика", 4, "Лекция", 90, "экран"])
    sheet.append([2, "Математика", 2, "Семинар", 90, "-"])
    sheet.append([3, "Информатика", 4, "Семинар", 90, "тех"])

    sheet = workbook["группы"]
    sheet.append([1, 1011, 25, "1, 2, 3"])
    sheet.append([2, 1012, 28, "1, 2"])


def create_template(output_file):
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = {
        "аудитории": ["id", "номер", "вместимость", "тип аудитории"],
        "преподаватели": ["id", "ФИО", "занятость в неделю", "преподает предметы (id)"],
        "предметы": [
            "id",
            "название",
            "часы в неделю",
            "тип занятия",
            "длительность",
            "дополнительны условия",
        ],
        "группы": ["id", "номер", "кол-во человек", "предметы (id)"],
    }

    for sheet_name, headers in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        write_headers(sheet, headers)

    workbook.save(output_file)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Конвертирует Excel-файл с данными расписания в JSON."
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=str(DEFAULT_INPUT_FILE),
        help=f"Excel-файл для чтения. По умолчанию: {DEFAULT_INPUT_FILE}",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=str(DEFAULT_OUTPUT_FILE),
        help=f"JSON-файл для записи. По умолчанию: {DEFAULT_OUTPUT_FILE}",
    )
    parser.add_argument(
        "--template",
        action="store_true",
        help=f"Создать шаблон Excel: {DEFAULT_TEMPLATE_FILE}",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    if args.template:
        create_template(DEFAULT_TEMPLATE_FILE)
        print(f"Шаблон Excel создан: {DEFAULT_TEMPLATE_FILE}")
        return

    input_file = Path(args.input)
    output_file = Path(args.output)

    if not input_file.exists():
        raise FileNotFoundError(
            f"Файл {input_file} не найден. "
            "Создайте шаблон командой: python excel_to_json.py --template"
        )

    data = convert_excel_to_json(input_file, output_file)
    print(f"JSON-файл создан: {output_file}")
    print(f"Разделы: {', '.join(data.keys())}")


if __name__ == "__main__":
    main()
